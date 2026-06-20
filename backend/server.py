from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Request
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timedelta, timezone
import jwt
from passlib.context import CryptContext

from pymongo import ReturnDocument
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import asyncio
import json

import gdrive

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ.get('MONGO_URL', 'mongodb+srv://KambojPress:Saaa60086009@kambojpressapp.ltwgpub.mongodb.net/?appName=KambojPressApp')
client = AsyncIOMotorClient(mongo_url)
db = client['KambojPressApp']

SECRET_KEY = os.environ.get('JWT_SECRET', 'press-order-book-secret-key-change-in-prod')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 30

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

app = FastAPI()
api = APIRouter(prefix="/api")

BUILTIN_STATUSES = ["Pending", "Designing", "Offset", "Digital Printing", "Printing", "Binding", "Ready", "Delivered"]
PROTECTED_STATUSES = {"Pending", "Delivered"}  # cannot be deleted/renamed

async def get_allowed_statuses() -> list:
    s = await db.settings.find_one({"id": "default"}, {"custom_statuses": 1})
    custom = (s or {}).get("custom_statuses") or []
    seen = set()
    out = []
    for st in list(BUILTIN_STATUSES) + list(custom):
        st = (st or "").strip()
        if st and st not in seen:
            seen.add(st)
            out.append(st)
    return out

# ---------- Models ----------
class UserCreate(BaseModel):
    name: str
    username: str
    password: str
    role: Literal["admin", "staff"] = "staff"

class UserOut(BaseModel):
    id: str
    name: str
    username: str
    role: str
    created_at: datetime

class LoginIn(BaseModel):
    username: str
    password: str

class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserOut

class ProductIn(BaseModel):
    name: str
    quantity: int = 1
    price: float = 0.0
    notes: Optional[str] = ""
    status: str = "Pending"

class ProductOut(ProductIn):
    id: str
    updated_at: Optional[datetime] = None

class OrderCreate(BaseModel):
    customer_name: str
    customer_phone: str
    customer_address: Optional[str] = ""
    products: List[ProductIn]
    assigned_user_ids: List[str] = []
    advance_paid: float = 0.0
    notes: Optional[str] = ""

class PaymentIn(BaseModel):
    amount: float
    note: Optional[str] = ""

class StatusUpdateIn(BaseModel):
    status: str

class OrderUpdateIn(BaseModel):
    assigned_user_ids: Optional[List[str]] = None
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_address: Optional[str] = None
    notes: Optional[str] = None
    advance_paid: Optional[float] = None
    products: Optional[List[ProductIn]] = None

class SettingsIn(BaseModel):
    app_name: Optional[str] = None
    company_name: Optional[str] = None
    company_phone: Optional[str] = None
    company_address: Optional[str] = None
    logo_base64: Optional[str] = None
    reminder_in_process_days: Optional[int] = None
    reminder_delivery_days: Optional[int] = None
    reminder_payment_days: Optional[int] = None
    custom_reminders: Optional[List[dict]] = None
    invoice_terms: Optional[str] = None
    custom_statuses: Optional[List[str]] = None

class DismissIn(BaseModel):
    key: str

# ---------- Helpers ----------
def hash_pw(p): return pwd_ctx.hash(p)
def verify_pw(p, h): return pwd_ctx.verify(p, h)

def make_token(user_id: str):
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(token: str = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        uid = payload.get("sub")
    except jwt.PyJWTError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": uid})
    if not user:
        raise HTTPException(401, "User not found")
    return user

async def require_admin(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin access required")
    return user

def user_public(u):
    return normalize_dates({"id": u["id"], "name": u["name"], "username": u["username"], "role": u["role"], "created_at": u["created_at"]})

def clean_doc(d):
    if not d: return d
    d.pop("_id", None)
    return d

def _iso_utc(v):
    if isinstance(v, datetime):
        if v.tzinfo is None:
            v = v.replace(tzinfo=timezone.utc)
        return v.isoformat()
    return v

def normalize_dates(obj):
    if isinstance(obj, dict):
        return {k: normalize_dates(_iso_utc(v) if isinstance(v, datetime) else v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [normalize_dates(x) for x in obj]
    if isinstance(obj, datetime):
        return _iso_utc(obj)
    return obj

# ---------- Auth ----------
@app.post("/auth/login")
@api.post("/auth/login", response_model=TokenOut)
async def login(data: LoginIn):
    user = await db.users.find_one({"username": data.username.lower().strip()})
    if not user or not verify_pw(data.password, user["password_hash"]):
        raise HTTPException(401, "Invalid username or password")
    return {"access_token": make_token(user["id"]), "token_type": "bearer", "user": user_public(user)}

@api.get("/auth/me", response_model=UserOut)
async def me(user=Depends(get_current_user)):
    return user_public(user)

@api.post("/users", response_model=UserOut)
async def create_user(data: UserCreate, admin=Depends(require_admin)):
    uname = data.username.lower().strip()
    if await db.users.find_one({"username": uname}):
        raise HTTPException(400, "Username already exists")
    user = {
        "id": str(uuid.uuid4()),
        "name": data.name.strip(),
        "username": uname,
        "password_hash": hash_pw(data.password),
        "role": data.role,
        "created_at": datetime.now(timezone.utc),
    }
    await db.users.insert_one(user)
    return user_public(user)

@api.get("/users", response_model=List[UserOut])
async def list_users(user=Depends(get_current_user)):
    users = await db.users.find().sort("created_at", -1).to_list(500)
    return [user_public(u) for u in users]

@api.delete("/users/{user_id}")
async def delete_user(user_id: str, admin=Depends(require_admin)):
    if user_id == admin["id"]:
        raise HTTPException(400, "Cannot delete yourself")
    res = await db.users.delete_one({"id": user_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}

# ---------- Customers ----------
async def upsert_customer(name: str, phone: str, address: str = ""):
    phone = phone.strip()
    name = name.strip()
    existing = await db.customers.find_one({"phone": phone})
    if existing:
        update = {"name": name or existing.get("name"), "updated_at": datetime.now(timezone.utc)}
        if address: update["address"] = address
        await db.customers.update_one({"id": existing["id"]}, {"$set": update})
        return existing["id"]
    cid = str(uuid.uuid4())
    await db.customers.insert_one({
        "id": cid, "name": name, "phone": phone, "address": address,
        "created_at": datetime.now(timezone.utc), "updated_at": datetime.now(timezone.utc),
    })
    return cid

@api.get("/customers")
async def list_customers(q: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if q:
        query = {"$or": [
            {"name": {"$regex": q, "$options": "i"}},
            {"phone": {"$regex": q, "$options": "i"}},
        ]}
    customers = await db.customers.find(query).sort("updated_at", -1).to_list(500)

    # Aggregate balance + order count per customer in a single pass.
    customer_ids = [c.get("id") for c in customers if c.get("id")]
    balance_map = {cid: {"total_balance": 0.0, "total_orders": 0, "total_business": 0.0} for cid in customer_ids}
    if customer_ids:
        async for o in db.orders.find({"customer_id": {"$in": customer_ids}}):
            cid = o.get("customer_id")
            if cid not in balance_map:
                continue
            total = sum((p.get("price", 0) or 0) for p in o.get("products", []))
            paid = (o.get("advance_paid", 0) or 0) + sum(p.get("amount", 0) for p in o.get("payments", []))
            balance = max(total - paid, 0)
            balance_map[cid]["total_balance"] += balance
            balance_map[cid]["total_business"] += total
            balance_map[cid]["total_orders"] += 1

    out = []
    for c in customers:
        d = normalize_dates(clean_doc(c))
        info = balance_map.get(d.get("id"), {"total_balance": 0.0, "total_orders": 0, "total_business": 0.0})
        d["total_balance"] = round(info["total_balance"], 2)
        d["total_business"] = round(info["total_business"], 2)
        d["total_orders"] = info["total_orders"]
        out.append(d)
    return out

@api.get("/customers/{customer_id}")
async def get_customer(customer_id: str, user=Depends(get_current_user)):
    c = await db.customers.find_one({"id": customer_id})
    if not c: raise HTTPException(404, "Customer not found")
    orders = await db.orders.find({"customer_id": customer_id}).sort("created_at", -1).to_list(500)
    return {"customer": normalize_dates(clean_doc(c)), "orders": [serialize_order(o) for o in orders]}

@api.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, admin=Depends(require_admin)):
    c = await db.customers.find_one({"id": customer_id})
    if not c: raise HTTPException(404, "Customer not found")
    order_count = await db.orders.count_documents({"customer_id": customer_id})
    if order_count > 0:
        raise HTTPException(400, f"Cannot delete customer with {order_count} order(s). Delete the orders first.")
    await db.customers.delete_one({"id": customer_id})
    return {"ok": True}

# ---------- Orders ----------
async def next_order_no():
    res = await db.counters.find_one_and_update(
        {"_id": "order_no"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    seq = res.get("seq", 1) if res else 1
    return f"ORD-{seq:04d}"

def order_total(order):
    return sum((p.get("price", 0) or 0) for p in order.get("products", []))

def order_paid(order):
    return (order.get("advance_paid", 0) or 0) + sum(p.get("amount", 0) for p in order.get("payments", []))

def serialize_order(o):
    o = clean_doc(o)
    total = order_total(o)
    paid = order_paid(o)
    o["total"] = total
    o["paid"] = paid
    o["balance"] = max(total - paid, 0)
    return normalize_dates(o)

@api.post("/orders")
async def create_order(data: OrderCreate, user=Depends(get_current_user)):
    if not data.customer_name.strip() or not data.customer_phone.strip():
        raise HTTPException(400, "Customer name and phone are required")
    if not data.products:
        raise HTTPException(400, "At least one product is required")
    customer_id = await upsert_customer(data.customer_name, data.customer_phone, data.customer_address or "")
    order_no = await next_order_no()
    now = datetime.now(timezone.utc)
    allowed_statuses = await get_allowed_statuses()
    products = []
    for p in data.products:
        st = p.status if p.status in allowed_statuses else "Pending"
        products.append({
            "id": str(uuid.uuid4()),
            "name": p.name.strip(),
            "quantity": p.quantity,
            "price": p.price,
            "notes": p.notes or "",
            "status": st,
            "updated_at": now,
        })
    assigned = list(data.assigned_user_ids or [])
    if user.get("role") == "staff" and user["id"] not in assigned:
        assigned.append(user["id"])
    order = {
        "id": str(uuid.uuid4()),
        "order_no": order_no,
        "customer_id": customer_id,
        "customer_name": data.customer_name.strip(),
        "customer_phone": data.customer_phone.strip(),
        "customer_address": data.customer_address or "",
        "products": products,
        "assigned_user_ids": assigned,
        "advance_paid": data.advance_paid or 0.0,
        "payments": [],
        "notes": data.notes or "",
        "created_by": user["id"],
        "created_by_name": user["name"],
        "created_at": now,
        "updated_at": now,
    }
    await db.orders.insert_one(order)
    schedule_gdrive_sync()
    return serialize_order(order)

@api.get("/orders")
async def list_orders(user=Depends(get_current_user)):
    orders = await db.orders.find().sort("created_at", -1).to_list(1000)
    return [serialize_order(o) for o in orders]

@api.get("/orders/balance/list")
async def list_balance_orders(user=Depends(get_current_user)):
    """Return only orders that still have an outstanding balance > 0,
    sorted by most outstanding first."""
    orders = await db.orders.find().sort("created_at", -1).to_list(2000)
    serialized = [serialize_order(o) for o in orders]
    pending = [o for o in serialized if (o.get("balance") or 0) > 0]
    pending.sort(key=lambda x: x.get("balance", 0), reverse=True)
    return pending

@api.get("/orders/{order_id}")
async def get_order(order_id: str, user=Depends(get_current_user)):
    o = await db.orders.find_one({"id": order_id})
    if not o: raise HTTPException(404, "Order not found")
    return serialize_order(o)

@api.patch("/orders/{order_id}")
async def update_order(order_id: str, data: OrderUpdateIn, user=Depends(get_current_user)):
    o = await db.orders.find_one({"id": order_id})
    if not o: raise HTTPException(404, "Order not found")
    now = datetime.now(timezone.utc)
    update = {"updated_at": now}
    if data.assigned_user_ids is not None: update["assigned_user_ids"] = data.assigned_user_ids
    if data.customer_address is not None: update["customer_address"] = data.customer_address
    if data.notes is not None: update["notes"] = data.notes
    # Staff is NOT allowed to change advance_paid after order creation (only admin).
    if data.advance_paid is not None and user.get("role") == "admin":
        update["advance_paid"] = float(data.advance_paid)

    new_name = data.customer_name.strip() if data.customer_name else None
    new_phone = data.customer_phone.strip() if data.customer_phone else None
    if new_name or new_phone:
        final_name = new_name or o.get("customer_name")
        final_phone = new_phone or o.get("customer_phone")
        if not final_name or not final_phone:
            raise HTTPException(400, "Customer name and phone are required")
        addr = data.customer_address if data.customer_address is not None else o.get("customer_address", "")
        new_customer_id = await upsert_customer(final_name, final_phone, addr)
        update["customer_id"] = new_customer_id
        update["customer_name"] = final_name
        update["customer_phone"] = final_phone

    if data.products is not None:
        if not data.products:
            raise HTTPException(400, "At least one product is required")
        allowed_statuses = await get_allowed_statuses()
        new_products = []
        for p in data.products:
            st = p.status if p.status in allowed_statuses else "Pending"
            prod = {
                "id": str(uuid.uuid4()),
                "name": p.name.strip(),
                "quantity": p.quantity,
                "price": p.price,
                "notes": p.notes or "",
                "status": st,
                "updated_at": now,
            }
            new_products.append(prod)
        update["products"] = new_products

    await db.orders.update_one({"id": order_id}, {"$set": update})
    o = await db.orders.find_one({"id": order_id})
    schedule_gdrive_sync()
    return serialize_order(o)

@api.patch("/orders/{order_id}/products/{product_id}")
async def update_product_status(order_id: str, product_id: str, data: StatusUpdateIn, user=Depends(get_current_user)):
    allowed_statuses = await get_allowed_statuses()
    if data.status not in allowed_statuses:
        raise HTTPException(400, "Invalid status")
    o = await db.orders.find_one({"id": order_id})
    if not o: raise HTTPException(404, "Order not found")
    if user["role"] == "staff" and user["id"] not in o.get("assigned_user_ids", []):
        raise HTTPException(403, "Not assigned to this order")
    now = datetime.now(timezone.utc)
    res = await db.orders.update_one(
        {"id": order_id, "products.id": product_id},
        {"$set": {"products.$.status": data.status, "products.$.updated_at": now, "updated_at": now}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Product not found in order")
    o = await db.orders.find_one({"id": order_id})
    schedule_gdrive_sync()
    return serialize_order(o)

@api.post("/orders/{order_id}/payments")
async def add_payment(order_id: str, data: PaymentIn, user=Depends(get_current_user)):
    if user["role"] == "staff":
        raise HTTPException(403, "Only admins can record payments")
    o = await db.orders.find_one({"id": order_id})
    if not o: raise HTTPException(404, "Order not found")
    payment = {"id": str(uuid.uuid4()), "amount": data.amount, "note": data.note or "", "at": datetime.now(timezone.utc), "by": user["name"]}
    await db.orders.update_one({"id": order_id}, {"$push": {"payments": payment}, "$set": {"updated_at": datetime.now(timezone.utc)}})
    o = await db.orders.find_one({"id": order_id})
    schedule_gdrive_sync()
    return serialize_order(o)


class PaymentUpdateIn(BaseModel):
    amount: Optional[float] = None
    note: Optional[str] = None


@api.patch("/orders/{order_id}/payments/{payment_id}")
async def update_payment(order_id: str, payment_id: str, data: PaymentUpdateIn, admin=Depends(require_admin)):
    o = await db.orders.find_one({"id": order_id})
    if not o: raise HTTPException(404, "Order not found")
    if not any(p.get("id") == payment_id for p in o.get("payments", [])):
        raise HTTPException(404, "Payment not found")
    update = {"updated_at": datetime.now(timezone.utc)}
    set_ops = {}
    if data.amount is not None:
        if data.amount < 0: raise HTTPException(400, "Amount must be >= 0")
        set_ops["payments.$[p].amount"] = float(data.amount)
    if data.note is not None:
        set_ops["payments.$[p].note"] = data.note
    if set_ops:
        await db.orders.update_one(
            {"id": order_id},
            {"$set": {**set_ops, **update}},
            array_filters=[{"p.id": payment_id}],
        )
    o = await db.orders.find_one({"id": order_id})
    schedule_gdrive_sync()
    return serialize_order(o)


@api.delete("/orders/{order_id}/payments/{payment_id}")
async def delete_payment(order_id: str, payment_id: str, admin=Depends(require_admin)):
    o = await db.orders.find_one({"id": order_id})
    if not o: raise HTTPException(404, "Order not found")
    res = await db.orders.update_one(
        {"id": order_id},
        {"$pull": {"payments": {"id": payment_id}}, "$set": {"updated_at": datetime.now(timezone.utc)}},
    )
    if res.modified_count == 0:
        raise HTTPException(404, "Payment not found")
    o = await db.orders.find_one({"id": order_id})
    schedule_gdrive_sync()
    return serialize_order(o)


@api.delete("/orders/{order_id}")
async def delete_order(order_id: str, admin=Depends(require_admin)):
    res = await db.orders.delete_one({"id": order_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Order not found")
    schedule_gdrive_sync()
    return {"ok": True}

# ---------- Reminders ----------
@api.get("/reminders")
async def reminders(request: Request, include_seen: bool = True, user=Depends(get_current_user)):
    # Header se pata chalega ki request kis page se aa rahi hai
    referer = request.headers.get("referer", "")
    
    # 🔥 Agar url mein 'dashboard' word hai, toh reminders mat dikhao!
    if "dashboard" in referer.lower() and user.get("role") != "staff":
        return []

    s = await db.settings.find_one({"id": "default"}) or {}
    in_proc_days = s.get("reminder_in_process_days", 2)
    delivery_days = s.get("reminder_delivery_days", 7)
    payment_days = s.get("reminder_payment_days", 10)
    customs = s.get("custom_reminders", [])
    dismissed = set(s.get("dismissed_reminders", []))

    now = datetime.now(timezone.utc)
    orders = await db.orders.find().sort("created_at", -1).to_list(1000)
    out = []
    
    for o in orders:
        try:
            statuses = [p.get("status") for p in o.get("products", [])]
            all_pending = all(st == "Pending" for st in statuses) if statuses else False
            all_delivered = all(st == "Delivered" for st in statuses) if statuses else False
            
            created_at = o.get("created_at")
            if not created_at: continue
            
            if isinstance(created_at, str):
                try:
                    if "Z" in created_at:
                        created_at = created_at.replace("Z", "+00:00")
                    created_at_dt = datetime.fromisoformat(created_at)
                except Exception:
                    continue
            else:
                created_at_dt = created_at
                if created_at_dt.tzinfo is None:
                    created_at_dt = created_at_dt.replace(tzinfo=timezone.utc)
            
            age_days = (now - created_at_dt).days
            
            total = sum((float(p.get("price") or 0)) for p in o.get("products", []))
            advance = float(o.get("advance_paid") or 0)
            p_payments = sum(float(p.get("amount") or 0) for p in o.get("payments", []))
            paid = advance + p_payments
            balance = max(total - paid, 0)

            def make(rtype, message, key_suffix=""):
                key = f"{o['id']}:{rtype}{(':' + key_suffix) if key_suffix else ''}"
                return {
                    "key": key,
                    "type": rtype,
                    "order_id": o["id"],
                    "order_no": o.get("order_no"),
                    "customer_name": o.get("customer_name"),
                    "customer_phone": o.get("customer_phone"),
                    "message": message,
                    "age_days": age_days,
                    "seen": key in dismissed,
                }

            if all_pending and age_days >= in_proc_days:
                out.append(make("in_process", f"Order not moved to In Process for {age_days} day(s)"))
            if not all_delivered and statuses and age_days >= delivery_days:
                out.append(make("delivery", f"Order not delivered after {age_days} day(s)"))
            if balance > 0 and age_days >= payment_days:
                out.append(make("payment", f"Payment pending: balance ₹{balance:.0f} after {age_days} day(s)"))
            
            for cr in customs:
                try:
                    d = int(cr.get("days", 0))
                    label = cr.get("label", "Custom reminder")
                    if d > 0 and age_days >= d and not all_delivered:
                        out.append(make("custom", f"{label} ({age_days} day(s) since creation)", key_suffix=label))
                except Exception:
                    pass
        except Exception as e:
            continue

    if not include_seen:
        out = [r for r in out if not r["seen"]]
        
    out.sort(key=lambda r: r.get("age_days", 0), reverse=True)
    return out
    
# ---------- Settings ----------
DEFAULT_SETTINGS = {
    "id": "default",
    "app_name": "Press Order Book",
    "company_name": "Press Order Book",
    "company_phone": "",
    "company_address": "",
    "logo_base64": "",
    "reminder_in_process_days": 2,
    "reminder_delivery_days": 7,
    "reminder_payment_days": 10,
    "custom_reminders": [],
    "invoice_terms": "1. Goods once sold will not be taken back.\n2. Payment due within 7 days of delivery.\n3. Subject to local jurisdiction.",
    "dismissed_reminders": [],
    "custom_statuses": [],
}

@api.get("/settings")
async def get_settings(user=Depends(get_current_user)):
    s = await db.settings.find_one({"id": "default"})
    if not s:
        s = DEFAULT_SETTINGS.copy()
        await db.settings.insert_one(s)
    return clean_doc(s)

@api.patch("/settings")
async def update_settings(data: SettingsIn, admin=Depends(require_admin)):
    s = await db.settings.find_one({"id": "default"})
    if not s:
        s = DEFAULT_SETTINGS.copy()
        await db.settings.insert_one(s)
    update = {k: v for k, v in data.dict().items() if v is not None}
    if update:
        await db.settings.update_one({"id": "default"}, {"$set": update})
    s = await db.settings.find_one({"id": "default"})
    return clean_doc(s)

# ---------- Branding (public — used by login screen and header) ----------
@app.get("/branding")
@api.get("/branding")
async def get_branding():
    s = await db.settings.find_one({"id": "default"})
    if not s:
        s = DEFAULT_SETTINGS.copy()
    return {
        "app_name": s.get("app_name") or "Press Order Book",
        "company_name": s.get("company_name") or "",
        "logo_base64": s.get("logo_base64") or "",
    }

# ---------- Backup (Excel) ----------
def _fmt_dt(val):
    if not val:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def _style_header(ws, num_cols):
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="F97316")  # orange-500
    align = Alignment(horizontal="left", vertical="center")
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = align
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except AttributeError:
            continue
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 50)


async def _order_summary_rows(include_serial: bool = True) -> list:
    """Build the compact 'order summary' table as a list-of-lists (header + rows).

    Layout (matches the user's screenshot): one row per product. The order-level
    fields (#, OrderNo, OrderDate, CustomerName, Phone, Advance, Balance) are
    only filled on the first product row of each order. Per-product fields
    (Qty, Product, Price, Total = Price*Qty) appear on every row.
    """
    orders = await db.orders.find().sort("created_at", 1).to_list(100000)
    header = [
        "#", "Order No", "Order Date", "Customer Name", "Phone",
        "Qty", "Product", "Price", "Total", "Advance", "Balance",
    ]
    if not include_serial:
        header = header[1:]
    rows = [header]

    ist_offset = timedelta(hours=5, minutes=30)
    serial = 0
    for o in orders:
        products = o.get("products") or []
        if not products:
            continue
        serial += 1
        created = o.get("created_at")
        if isinstance(created, str):
            try:
                created = datetime.fromisoformat(created.replace("Z", "+00:00"))
            except Exception:
                created = None
        if isinstance(created, datetime):
            if created.tzinfo is None:
                created = created.replace(tzinfo=timezone.utc)
            local = created.astimezone(timezone(ist_offset))
            order_date = local.strftime("%d-%m-%Y %I:%M %p")
        else:
            order_date = ""

        order_total = sum(float(p.get("price") or 0) for p in products)
        advance = float(o.get("advance_paid") or 0)
        payments_sum = sum(float(p.get("amount") or 0) for p in (o.get("payments") or []))
        paid = advance + payments_sum
        balance = max(order_total - paid, 0)

        for idx, p in enumerate(products):
            qty = float(p.get("quantity") or 0)
            line_total = float(p.get("price") or 0)
            unit_price = (line_total / qty) if qty else line_total
            if idx == 0:
                row = [
                    serial,
                    o.get("order_no") or "",
                    order_date,
                    o.get("customer_name") or "",
                    o.get("customer_phone") or "",
                    qty,
                    p.get("name") or "",
                    round(unit_price, 2),
                    round(line_total, 2),
                    round(advance, 2) if advance else "",
                    round(balance, 2) if balance else "",
                ]
            else:
                row = [
                    "", "", "", "", "",
                    qty,
                    p.get("name") or "",
                    round(unit_price, 2),
                    round(line_total, 2),
                    "", "",
                ]
            if not include_serial:
                row = row[1:]
            rows.append(row)
    return rows


async def _build_order_summary_workbook() -> bytes:
    rows = await _order_summary_rows(include_serial=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Summary"
    for r in rows:
        ws.append(r)
    # Style header
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = center
    # Auto width
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _build_backup_workbook() -> bytes:
    users = await db.users.find().to_list(10000)
    customers = await db.customers.find().to_list(10000)
    orders = await db.orders.find().sort("created_at", -1).to_list(100000)
    settings_doc = await db.settings.find_one({"id": "default"}) or {}

    customer_map = {c.get("id"): c for c in customers}

    wb = Workbook()

    # Sheet 1 - Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    ws.append(["App Name", settings_doc.get("app_name") or "Press Order Book"])
    ws.append(["Company Name", settings_doc.get("company_name") or ""])
    ws.append(["Company Phone", settings_doc.get("company_phone") or ""])
    ws.append(["Company Address", settings_doc.get("company_address") or ""])
    ws.append(["Exported At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws.append(["Total Users", len(users)])
    ws.append(["Total Customers", len(customers)])
    ws.append(["Total Orders", len(orders)])
    total_revenue = 0.0
    total_paid_all = 0.0
    total_balance_all = 0.0
    for o in orders:
        t = sum((p.get("price", 0) or 0) for p in o.get("products", []))
        pd = (o.get("advance_paid", 0) or 0) + sum(p.get("amount", 0) for p in o.get("payments", []))
        total_revenue += t
        total_paid_all += pd
        total_balance_all += max(t - pd, 0)
    ws.append(["Total Order Value (INR)", round(total_revenue, 2)])
    ws.append(["Total Paid (INR)", round(total_paid_all, 2)])
    ws.append(["Total Outstanding Balance (INR)", round(total_balance_all, 2)])
    _style_header(ws, 2)
    _autosize(ws)

    # Sheet 2 - Orders (one row per order)
    ws = wb.create_sheet("Orders")
    headers = [
        "Order No", "Date Created", "Last Updated", "Customer Name", "Customer Phone", "Customer Address",
        "Products Count", "Products (name x qty @ price [status])",
        "Total (INR)", "Advance Paid (INR)", "Other Payments (INR)", "Paid (INR)", "Balance (INR)",
        "Created By", "Assigned User IDs", "Order Notes",
    ]
    ws.append(headers)
    for o in orders:
        prods = o.get("products", [])
        total = sum((p.get("price", 0) or 0) for p in prods)
        advance = o.get("advance_paid", 0) or 0
        other_pmts = sum(p.get("amount", 0) for p in o.get("payments", []))
        paid = advance + other_pmts
        balance = max(total - paid, 0)
        product_str = "; ".join(
            f"{p.get('name','')} x{p.get('quantity',0)} @ {p.get('price',0)} [{p.get('status','')}]"
            for p in prods
        )
        ws.append([
            o.get("order_no", ""),
            _fmt_dt(o.get("created_at")),
            _fmt_dt(o.get("updated_at")),
            o.get("customer_name", ""),
            o.get("customer_phone", ""),
            o.get("customer_address", ""),
            len(prods),
            product_str,
            round(total, 2),
            round(advance, 2),
            round(other_pmts, 2),
            round(paid, 2),
            round(balance, 2),
            o.get("created_by_name", ""),
            ", ".join(o.get("assigned_user_ids", []) or []),
            o.get("notes", ""),
        ])
    _style_header(ws, len(headers))
    _autosize(ws)

    # Sheet 3 - Order Products (one row per product line item)
    ws = wb.create_sheet("Order Products")
    headers = [
        "Order No", "Order Date", "Customer Name", "Customer Phone",
        "Product Name", "Quantity", "Price (INR)", "Status", "Product Notes", "Product Updated At",
    ]
    ws.append(headers)
    for o in orders:
        for p in o.get("products", []):
            ws.append([
                o.get("order_no", ""),
                _fmt_dt(o.get("created_at")),
                o.get("customer_name", ""),
                o.get("customer_phone", ""),
                p.get("name", ""),
                p.get("quantity", 0),
                round(p.get("price", 0) or 0, 2),
                p.get("status", ""),
                p.get("notes", ""),
                _fmt_dt(p.get("updated_at")),
            ])
    _style_header(ws, len(headers))
    _autosize(ws)

    # Sheet 4 - Payments
    ws = wb.create_sheet("Payments")
    headers = ["Order No", "Customer Name", "Type", "Amount (INR)", "Date", "Recorded By", "Note"]
    ws.append(headers)
    for o in orders:
        if (o.get("advance_paid") or 0) > 0:
            ws.append([
                o.get("order_no", ""),
                o.get("customer_name", ""),
                "Advance",
                round(o.get("advance_paid", 0) or 0, 2),
                _fmt_dt(o.get("created_at")),
                o.get("created_by_name", ""),
                "Initial advance at order creation",
            ])
        for pmt in o.get("payments", []):
            ws.append([
                o.get("order_no", ""),
                o.get("customer_name", ""),
                "Payment",
                round(pmt.get("amount", 0) or 0, 2),
                _fmt_dt(pmt.get("at")),
                pmt.get("by", ""),
                pmt.get("note", ""),
            ])
    _style_header(ws, len(headers))
    _autosize(ws)

    # Sheet 5 - Customers (with balance)
    ws = wb.create_sheet("Customers")
    headers = ["Name", "Phone", "Address", "Orders", "Total Business (INR)", "Total Paid (INR)", "Outstanding Balance (INR)", "Created At", "Last Updated"]
    ws.append(headers)
    # Compute per-customer aggregates
    per_cust = {}
    for o in orders:
        cid = o.get("customer_id")
        if not cid:
            continue
        t = sum((p.get("price", 0) or 0) for p in o.get("products", []))
        pd = (o.get("advance_paid", 0) or 0) + sum(p.get("amount", 0) for p in o.get("payments", []))
        b = max(t - pd, 0)
        agg = per_cust.setdefault(cid, {"orders": 0, "biz": 0.0, "paid": 0.0, "bal": 0.0})
        agg["orders"] += 1
        agg["biz"] += t
        agg["paid"] += pd
        agg["bal"] += b
    for c in customers:
        cid = c.get("id")
        agg = per_cust.get(cid, {"orders": 0, "biz": 0.0, "paid": 0.0, "bal": 0.0})
        ws.append([
            c.get("name", ""),
            c.get("phone", ""),
            c.get("address", ""),
            agg["orders"],
            round(agg["biz"], 2),
            round(agg["paid"], 2),
            round(agg["bal"], 2),
            _fmt_dt(c.get("created_at")),
            _fmt_dt(c.get("updated_at")),
        ])
    _style_header(ws, len(headers))
    _autosize(ws)

    # Sheet 6 - Outstanding Balance (orders with balance > 0)
    ws = wb.create_sheet("Outstanding Balance")
    headers = ["Order No", "Date", "Customer Name", "Customer Phone", "Total (INR)", "Paid (INR)", "Balance (INR)", "Days Old"]
    ws.append(headers)
    now = datetime.now(timezone.utc)
    pending_rows = []
    for o in orders:
        t = sum((p.get("price", 0) or 0) for p in o.get("products", []))
        pd = (o.get("advance_paid", 0) or 0) + sum(p.get("amount", 0) for p in o.get("payments", []))
        b = max(t - pd, 0)
        if b <= 0:
            continue
        created = o.get("created_at")
        days_old = ""
        if isinstance(created, datetime):
            try:
                if created.tzinfo is None:
                    created = created.replace(tzinfo=timezone.utc)
                days_old = (now - created).days
            except Exception:
                days_old = ""
        pending_rows.append([
            o.get("order_no", ""), _fmt_dt(o.get("created_at")),
            o.get("customer_name", ""), o.get("customer_phone", ""),
            round(t, 2), round(pd, 2), round(b, 2), days_old,
        ])
    pending_rows.sort(key=lambda r: r[6], reverse=True)  # highest balance first
    for r in pending_rows:
        ws.append(r)
    _style_header(ws, len(headers))
    _autosize(ws)

    # Sheet 7 - Users
    ws = wb.create_sheet("Users")
    headers = ["Username", "Name", "Role", "Created At"]
    ws.append(headers)
    for u in users:
        ws.append([u.get("username", ""), u.get("name", ""), u.get("role", ""), _fmt_dt(u.get("created_at"))])
    _style_header(ws, len(headers))
    _autosize(ws)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


@app.get("/backup")
@api.get("/backup")
async def export_backup(admin=Depends(require_admin)):
    data = await _build_backup_workbook()
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"press-order-book-backup-{stamp}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@api.get("/backup/summary")
async def export_order_summary(admin=Depends(require_admin)):
    data = await _build_order_summary_workbook()
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"order-summary-{stamp}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ---------- Google Drive Sync ----------
_GDRIVE_SYNC_TASK: Optional[asyncio.Task] = None
_GDRIVE_SYNC_DEBOUNCE_SEC = 3.0


async def _load_gdrive_config() -> Optional[dict]:
    return await db.gdrive_config.find_one({"id": "default"})


async def _save_gdrive_config(update: dict) -> dict:
    doc = await db.gdrive_config.find_one_and_update(
        {"id": "default"},
        {"$set": update, "$setOnInsert": {"id": "default"}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return doc


def _public_gdrive_status(cfg: Optional[dict]) -> dict:
    if not cfg:
        return {"connected": False, "has_config": False}
    sa = cfg.get("service_account_json") or {}
    has_config = bool(cfg.get("service_account_json") or cfg.get("spreadsheet_id") or cfg.get("folder_id"))
    return {
        "connected": bool(cfg.get("service_account_json") and cfg.get("spreadsheet_id")),
        "has_config": has_config,
        "auto_sync": bool(cfg.get("auto_sync", True)),
        "service_account_email": sa.get("client_email"),
        "spreadsheet_id": cfg.get("spreadsheet_id"),
        "spreadsheet_name": cfg.get("spreadsheet_name"),
        "spreadsheet_url": gdrive.sheet_url(cfg["spreadsheet_id"]) if cfg.get("spreadsheet_id") else None,
        "last_sync_at": cfg.get("last_sync_at"),
        "last_sync_status": cfg.get("last_sync_status"),
        "last_sync_error": cfg.get("last_sync_error"),
    }


async def _do_gdrive_sync() -> dict:
    cfg = await _load_gdrive_config()
    if not cfg or not cfg.get("service_account_json") or not cfg.get("spreadsheet_id"):
        raise GDriveNotConfigured("Google Drive abhi tak connect nahi hua hai.")
    sa_info = cfg["service_account_json"]
    sheet_id = cfg["spreadsheet_id"]
    rows = await _order_summary_rows(include_serial=True)
    try:
        # Make sure SA still has edit access; raises if not.
        sheet_id = await asyncio.to_thread(gdrive.ensure_spreadsheet, sa_info, sheet_id)
        result = await asyncio.to_thread(gdrive.write_rows, sa_info, sheet_id, rows)
    except gdrive.GDriveError as e:
        await _save_gdrive_config({
            "last_sync_status": "error",
            "last_sync_error": str(e),
            "last_sync_at": datetime.now(timezone.utc).isoformat(),
        })
        raise
    await _save_gdrive_config({
        "last_sync_status": "ok",
        "last_sync_error": None,
        "last_sync_at": datetime.now(timezone.utc).isoformat(),
        "last_sync_rows": result.get("updated_rows", 0),
    })
    return {
        "spreadsheet_id": sheet_id,
        "spreadsheet_url": gdrive.sheet_url(sheet_id),
        "rows": result.get("updated_rows", 0),
    }


class GDriveNotConfigured(Exception):
    pass


async def _debounced_sync():
    try:
        await asyncio.sleep(_GDRIVE_SYNC_DEBOUNCE_SEC)
        await _do_gdrive_sync()
    except asyncio.CancelledError:
        pass
    except GDriveNotConfigured:
        pass
    except Exception as e:
        logging.warning(f"GDrive auto-sync failed: {e}")


def schedule_gdrive_sync() -> None:
    """Fire-and-forget: schedule a debounced sync. Safe to call from any
    order-mutating endpoint. If Drive isn't connected or auto-sync is off,
    the task is a cheap no-op."""
    global _GDRIVE_SYNC_TASK
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return

    async def _wrap():
        cfg = await _load_gdrive_config()
        if not cfg or not cfg.get("auto_sync", True):
            return
        if not cfg.get("service_account_json") or not cfg.get("spreadsheet_id"):
            return
        await _debounced_sync()

    if _GDRIVE_SYNC_TASK and not _GDRIVE_SYNC_TASK.done():
        _GDRIVE_SYNC_TASK.cancel()
    _GDRIVE_SYNC_TASK = loop.create_task(_wrap())


class GDriveConnectIn(BaseModel):
    service_account_json: dict
    spreadsheet: str  # URL or ID of an existing Google Sheet (user-created + shared with SA)
    auto_sync: bool = True


@api.get("/gdrive/status")
async def gdrive_status(admin=Depends(require_admin)):
    cfg = await _load_gdrive_config()
    return _public_gdrive_status(cfg)


@api.post("/gdrive/connect")
async def gdrive_connect(data: GDriveConnectIn, admin=Depends(require_admin)):
    sa = data.service_account_json
    if not isinstance(sa, dict) or sa.get("type") != "service_account" or not sa.get("client_email"):
        raise HTTPException(400, "Yeh service-account JSON valid nahi lag rahi. Google Cloud Console se naya key download karein.")
    spreadsheet_id = gdrive.parse_spreadsheet_id(data.spreadsheet)
    if not spreadsheet_id:
        raise HTTPException(400, "Google Sheet URL ya ID zaruri hai.")
    try:
        meta = await asyncio.to_thread(gdrive.verify_connection, sa, spreadsheet_id)
    except gdrive.GDriveError as e:
        raise HTTPException(400, str(e))
    # Wipe out any stale folder config from previous attempts.
    await db.gdrive_config.delete_one({"id": "default"})
    await _save_gdrive_config({
        "service_account_json": sa,
        "spreadsheet_id": spreadsheet_id,
        "spreadsheet_name": meta.get("name"),
        "auto_sync": data.auto_sync,
        "connected_at": datetime.now(timezone.utc).isoformat(),
    })
    # Run first sync immediately so the user sees data right away.
    try:
        sync_result = await _do_gdrive_sync()
    except Exception as e:
        raise HTTPException(400, f"Connected, but first sync failed: {e}")
    cfg = await _load_gdrive_config()
    out = _public_gdrive_status(cfg)
    out["first_sync"] = sync_result
    return out


@api.post("/gdrive/sync")
async def gdrive_sync_now(admin=Depends(require_admin)):
    try:
        result = await _do_gdrive_sync()
    except GDriveNotConfigured as e:
        raise HTTPException(400, str(e))
    except gdrive.GDriveError as e:
        raise HTTPException(400, str(e))
    return result


@api.patch("/gdrive/auto-sync")
async def gdrive_toggle_auto(data: dict, admin=Depends(require_admin)):
    await _save_gdrive_config({"auto_sync": bool(data.get("auto_sync", True))})
    cfg = await _load_gdrive_config()
    return _public_gdrive_status(cfg)


@api.delete("/gdrive/disconnect")
async def gdrive_disconnect(admin=Depends(require_admin)):
    await db.gdrive_config.delete_one({"id": "default"})
    return {"connected": False, "has_config": False}

# ---------- Stats ----------
@app.get("/stats")
@api.get("/stats")
async def stats(user=Depends(get_current_user), year: Optional[int] = None, month: Optional[int] = None):
    """
    All-time + today by default. Optional ?year=YYYY filters by year (IST).
    Optional ?year=YYYY&month=MM filters by specific month.
    """
    orders = await db.orders.find().to_list(5000)

    def aggregate(order_list):
        total_orders = len(order_list)
        total_revenue = sum((order_total(o) or 0) for o in order_list)
        total_paid = sum((order_paid(o) or 0) for o in order_list)
        balance = max(total_revenue - total_paid, 0)
        pending = in_progress = delivered = 0
        for o in order_list:
            statuses = [p.get("status") for p in o.get("products", [])]
            if not statuses: continue
            if all(s == "Delivered" for s in statuses):
                delivered += 1
            elif all(s == "Pending" for s in statuses):
                pending += 1
            else:
                in_progress += 1
        return {
            "total_orders": total_orders,
            "total_revenue": total_revenue,
            "total_paid": total_paid,
            "balance_due": balance,
            "pending": pending,
            "in_progress": in_progress,
            "delivered": delivered,
        }

    IST = timezone(timedelta(hours=5, minutes=30))

    def order_ist_date(o):
        ca = o.get("created_at")
        if not ca: return None
        if isinstance(ca, str):
            try:
                ca = datetime.fromisoformat(ca.replace("Z", "+00:00"))
            except Exception:
                return ca[:10]
        if ca.tzinfo is None:
            ca = ca.replace(tzinfo=timezone.utc)
        return ca.astimezone(IST)

    # Period filter (year / month)
    period_orders = orders
    period_label = None
    if year is not None:
        period_orders = []
        for o in orders:
            d = order_ist_date(o)
            if isinstance(d, datetime) and d.year == year:
                if month is None or d.month == month:
                    period_orders.append(o)
        if month is not None:
            period_label = f"{year}-{month:02d}"
        else:
            period_label = str(year)

    # Today aggregate
    today_str = datetime.now(IST).strftime("%Y-%m-%d")
    today_orders = []
    for o in orders:
        d = order_ist_date(o)
        ds = d.strftime("%Y-%m-%d") if isinstance(d, datetime) else (d if isinstance(d, str) else None)
        if ds == today_str:
            today_orders.append(o)

    result = aggregate(period_orders if year is not None else orders)
    result["today"] = aggregate(today_orders)
    result["period"] = period_label
    result["reminders"] = []
    return result


@api.get("/stats/periods")
async def stats_periods(admin=Depends(require_admin)):
    """Returns available years and (year, month) pairs based on existing orders (IST)."""
    orders = await db.orders.find({}, {"created_at": 1}).to_list(20000)
    IST = timezone(timedelta(hours=5, minutes=30))
    year_set = set()
    ym_set = set()
    for o in orders:
        ca = o.get("created_at")
        if not ca: continue
        if isinstance(ca, str):
            try:
                ca = datetime.fromisoformat(ca.replace("Z", "+00:00"))
            except Exception:
                continue
        if ca.tzinfo is None:
            ca = ca.replace(tzinfo=timezone.utc)
        d = ca.astimezone(IST)
        year_set.add(d.year)
        ym_set.add((d.year, d.month))
    return {
        "years": sorted(year_set, reverse=True),
        "months": sorted([{"year": y, "month": m} for (y, m) in ym_set], key=lambda x: (-x["year"], -x["month"])),
    }
# ---------- Statuses ----------
class StatusIn(BaseModel):
    name: str

@api.get("/statuses")
async def list_statuses(user=Depends(get_current_user)):
    statuses = await get_allowed_statuses()
    s = await db.settings.find_one({"id": "default"}, {"custom_statuses": 1}) or {}
    return {
        "statuses": statuses,
        "builtin": BUILTIN_STATUSES,
        "custom": s.get("custom_statuses") or [],
        "protected": list(PROTECTED_STATUSES),
    }

@api.post("/statuses")
async def add_status(data: StatusIn, admin=Depends(require_admin)):
    name = (data.name or "").strip()
    if not name:
        raise HTTPException(400, "Status name is required")
    if len(name) > 40:
        raise HTTPException(400, "Status name too long")
    existing = await get_allowed_statuses()
    if name in existing:
        raise HTTPException(400, "Status already exists")
    s = await db.settings.find_one({"id": "default"}) or {}
    custom = list(s.get("custom_statuses") or [])
    custom.append(name)
    await db.settings.update_one({"id": "default"}, {"$set": {"custom_statuses": custom}}, upsert=True)
    return {"ok": True, "statuses": await get_allowed_statuses()}

@api.delete("/statuses/{name}")
async def delete_status(name: str, admin=Depends(require_admin)):
    if name in PROTECTED_STATUSES:
        raise HTTPException(400, f"'{name}' is protected and cannot be deleted")
    if name in BUILTIN_STATUSES:
        raise HTTPException(400, f"'{name}' is a built-in status and cannot be deleted")
    s = await db.settings.find_one({"id": "default"}) or {}
    custom = [c for c in (s.get("custom_statuses") or []) if c != name]
    await db.settings.update_one({"id": "default"}, {"$set": {"custom_statuses": custom}}, upsert=True)
    # Reset any products still using this status back to "Pending"
    await db.orders.update_many(
        {"products.status": name},
        {"$set": {"products.$[p].status": "Pending"}},
        array_filters=[{"p.status": name}],
    )
    return {"ok": True, "statuses": await get_allowed_statuses()}

# ---------- Init ----------
async def seed_admin():
    if await db.users.count_documents({}) == 0:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "name": "Admin",
            "username": "admin",
            "password_hash": hash_pw("admin123"),
            "role": "admin",
            "created_at": datetime.now(timezone.utc),
        })
        logging.info("Seeded default admin: admin / admin123")
    if not await db.settings.find_one({"id": "default"}):
        await db.settings.insert_one(DEFAULT_SETTINGS.copy())
    else:
        await db.settings.update_one(
            {"id": "default", "app_name": {"$exists": False}},
            {"$set": {"app_name": "Press Order Book"}},
        )

async def migrate_statuses():
    res = await db.orders.update_many(
        {"products.status": "Printing"},
        {"$set": {"products.$[p].status": "Digital Printing"}},
        array_filters=[{"p.status": "Printing"}],
    )
    if res.modified_count:
        logging.info(f"Migrated {res.modified_count} order(s): Printing -> Digital Printing")

    # Migrate Cutting -> Binding
    res = await db.orders.update_many(
        {"products.status": "Cutting"},
        {"$set": {"products.$[p].status": "Binding"}},
        array_filters=[{"p.status": "Cutting"}],
    )
    if res.modified_count:
        logging.info(f"Migrated {res.modified_count} order(s): Cutting -> Binding")

    # Migrate Packing -> Pending (Flex removed from builtin statuses)
    res = await db.orders.update_many(
        {"products.status": "Packing"},
        {"$set": {"products.$[p].status": "Pending"}},
        array_filters=[{"p.status": "Packing"}],
    )
    if res.modified_count:
        logging.info(f"Migrated {res.modified_count} order(s): Packing -> Pending")

    # Flex removed: reset existing Flex products to Pending
    res = await db.orders.update_many(
        {"products.status": "Flex"},
        {"$set": {"products.$[p].status": "Pending"}},
        array_filters=[{"p.status": "Flex"}],
    )
    if res.modified_count:
        logging.info(f"Migrated {res.modified_count} order(s): Flex -> Pending")

    # Rename Screen Printing -> Printing
    res = await db.orders.update_many(
        {"products.status": "Screen Printing"},
        {"$set": {"products.$[p].status": "Printing"}},
        array_filters=[{"p.status": "Screen Printing"}],
    )
    if res.modified_count:
        logging.info(f"Migrated {res.modified_count} order(s): Screen Printing -> Printing")

@app.on_event("startup")
async def startup():
    await seed_admin()
    await migrate_statuses()
    # Performance: create indexes for hot query paths
    try:
        await db.users.create_index("username", unique=True)
        await db.users.create_index("id", unique=True)
        await db.orders.create_index("id", unique=True)
        await db.orders.create_index("created_at")
        await db.orders.create_index("customer_id")
        await db.orders.create_index("assigned_user_ids")
        await db.orders.create_index("order_no")
        await db.customers.create_index("id", unique=True)
        await db.customers.create_index("phone")
        await db.customers.create_index("updated_at")
    except Exception as e:
        logging.warning(f"Index creation warning: {e}")

@app.post("/auth/login")
async def bypass_login(data: LoginIn):
    return await login(data)

@app.get("/stats")
async def bypass_stats(user=Depends(get_current_user)):
    return await stats(user)

@app.get("/orders")
async def bypass_orders(user=Depends(get_current_user)):
    return await list_orders(user)

@app.post("/orders")
async def bypass_create_order(data: OrderCreate, user=Depends(get_current_user)):
    return await create_order(data, user)

@app.middleware("http")
async def add_api_prefix_if_missing(request, call_next):
    if request.url.path.startswith("/auth") or request.url.path.startswith("/orders") or request.url.path.startswith("/stats") or request.url.path.startswith("/customers") or request.url.path.startswith("/settings"):
        request.scope["path"] = f"/api{request.url.path}"
    return await call_next(request)

# --- YAHAN SE COPY KAREIN ---
# Reminders Dismiss aur Restore ka special bypass rule
@app.post("/reminders/dismiss")
@api.post("/reminders/dismiss")
async def bypass_dismiss_reminder(data: DismissIn):
    # Bina kisi strict admin check ke seedha database mein update
    await db.settings.update_one(
        {"id": "default"}, 
        {"$addToSet": {"dismissed_reminders": data.key}}
    )
    return {"ok": True}

@app.post("/reminders/restore")
@api.post("/reminders/restore")
async def bypass_restore_reminder(data: DismissIn):
    await db.settings.update_one(
        {"id": "default"}, 
        {"$pull": {"dismissed_reminders": data.key}}
    )
    return {"ok": True}
# --- COPY YAHAN KHATAM ---

app.include_router(api)

from starlette.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=500)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
